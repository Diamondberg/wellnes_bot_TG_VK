"""
Генерация отчётов: TXT и Excel.

Эти отчёты идут АДМИНУ (на email и/или в Telegram).
Клиенту в чат показывается короткая версия — она формируется в платформенном слое.

Формат TXT:
  - простой человекочитаемый текст
  - удобно открывать в любом редакторе
  - подходит для пересылки

Формат Excel:
  - удобен для CRM-систем (импорт)
  - визуально красив (заголовки, цвета)
  - содержит все данные пользователя + все ответы + результаты

Архитектура:
  - на вход: данные пользователя (dataclass) + результат теста + список ответов
  - на выход: bytes (для отправки/сохранения)
  - модуль НЕ знает о БД и платформах
"""

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import List, Optional

from core.questions import get_question, SYSTEMS
from core.test_engine import TestResult, AnswerInput, is_yes


# ════════════════════════════════════════════════════════════
#  Входные данные для отчёта
# ════════════════════════════════════════════════════════════
@dataclass(frozen=True)
class UserSnapshot:
    """
    Снимок данных пользователя для отчёта.

    Это «вырезка» из БД — всё, что нужно отчёту, в плоском виде.
    Чтобы reports.py не зависел от моделей SQLAlchemy.
    """
    lead_number: int
    full_name: str
    phone: str
    platform: str               # "telegram" / "vk" / "max"
    platform_user_id: int       # tg/vk/max user_id (текущая платформа)
    platform_username: Optional[str] = None  # @username, если есть

    # ─── ID самого лида во всех платформах (для кликабельных ссылок) ──
    # Заполняются из БД-записи юзера. У одного человека может быть
    # заполнено сразу несколько (если поженили вручную или прошёл тест
    # на нескольких платформах с одинаковым telegram_user_id).
    lead_tg_user_id: Optional[int] = None
    lead_tg_username: Optional[str] = None
    lead_vk_user_id: Optional[int] = None
    lead_max_user_id: Optional[int] = None

    # ─── Информация о реферере (если применимо) ──────────────
    # Все поля Optional — для обратной совместимости.
    # Если реферер = дистрибьютор → referrer_name=None, в письме покажется
    # пометка «🌐 Зашёл напрямую» (см. core/notifier.py).
    referrer_name: Optional[str] = None
    referrer_phone: Optional[str] = None

    # Дополнительные поля реферера для письма админу (Шаг 5.2).
    # Чтобы админ мог связаться с реферером по нужному каналу
    # и предложить участие в реферальной программе.
    referrer_email: Optional[str] = None
    referrer_tg_username: Optional[str] = None   # без @, например "Diamondberg"
    referrer_tg_user_id: Optional[int] = None    # для ссылки tg://user?id=...
    referrer_vk_user_id: Optional[int] = None    # для ссылки https://vk.com/id...
    referrer_max_user_id: Optional[int] = None   # для будущей ссылки в Max


# ════════════════════════════════════════════════════════════
#  Дисклеймер
# ════════════════════════════════════════════════════════════
DISCLAIMER = (
    "ВАЖНОЕ УВЕДОМЛЕНИЕ\n"
    "Данный тест носит исключительно информационный характер.\n"
    "Он не является медицинской консультацией, диагнозом или заменой визита к врачу.\n"
    "При наличии серьёзных симптомов обратитесь к квалифицированному специалисту."
)


# ════════════════════════════════════════════════════════════
#  Генерация TXT
# ════════════════════════════════════════════════════════════
def generate_txt_report(
    user: UserSnapshot,
    answers: List[AnswerInput],
    result: TestResult,
) -> bytes:
    """
    TXT-отчёт для админа. Возвращает bytes (UTF-8).
    """
    lines: List[str] = []

    # ─── Шапка ──────────────────────────────────────────────
    lines.append("═" * 70)
    lines.append(f"WELLNESS TEST — ЗАЯВКА #{user.lead_number}")
    lines.append("═" * 70)
    lines.append("")
    lines.append(f"Дата:       {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"ФИО:        {user.full_name}")
    lines.append(f"Телефон:    {user.phone}")
    lines.append(f"Платформа:  {user.platform}")
    lines.append(f"User ID:    {user.platform_user_id}")
    if user.platform_username:
        lines.append(f"Username:   @{user.platform_username}")

    # ─── Блок реферера (Шаг 5.2: расширенный или «холодный») ──
    if user.referrer_name:
        lines.append("")
        lines.append("─── Реферал от ───")
        lines.append(f"ФИО:        {user.referrer_name}")
        if user.referrer_phone:
            lines.append(f"Телефон:    {user.referrer_phone}")
        if user.referrer_email:
            lines.append(f"Email:      {user.referrer_email}")
        if user.referrer_tg_username:
            lines.append(f"Telegram:   @{user.referrer_tg_username}")
        elif user.referrer_tg_user_id:
            lines.append(f"Telegram:   id={user.referrer_tg_user_id}")
        if user.referrer_vk_user_id:
            lines.append(f"VK:         https://vk.com/id{user.referrer_vk_user_id}")
        if user.referrer_max_user_id:
            lines.append(f"Max ID:     {user.referrer_max_user_id}")
    else:
        lines.append("")
        lines.append("─── Реферал ───")
        lines.append("🌐 Зашёл напрямую (без реф-ссылки)")

    # ─── Сводка по системам ────────────────────────────────
    lines.append("")
    lines.append("═" * 70)
    lines.append("РЕЗУЛЬТАТЫ ПО СИСТЕМАМ ОРГАНИЗМА")
    lines.append("═" * 70)
    lines.append("")

    for sys_result in result.systems:
        marker = sys_result.status.emoji
        lines.append(
            f"{marker} {sys_result.system_name}: "
            f"{sys_result.score}/{sys_result.max_score} "
            f"({sys_result.percentage:.0f}%) — {sys_result.status.label_ru}"
        )
        lines.append(f"   {sys_result.recommendation}")
        lines.append("")

    # ─── Сегментация лида (для админа) ─────────────────────
    lines.append("─── Сегментация лида ───")
    if result.critical_count >= 3:
        segment = "🔥 ГОРЯЧИЙ ЛИД (3+ систем в критическом состоянии)"
    elif result.critical_count >= 1:
        segment = "🟠 ТЁПЛЫЙ ЛИД (есть критические системы)"
    elif result.warning_count >= 3:
        segment = "🟡 ВНИМАТЕЛЬНЫЙ ЛИД (множественные warning'и)"
    else:
        segment = "🟢 ХОЛОДНЫЙ ЛИД (всё неплохо)"
    lines.append(segment)
    lines.append(f"Критических: {result.critical_count}, Warning'ов: {result.warning_count}")

    # ─── Детальные ответы ──────────────────────────────────
    lines.append("")
    lines.append("═" * 70)
    lines.append("ДЕТАЛЬНЫЕ ОТВЕТЫ")
    lines.append("═" * 70)
    lines.append("")

    for ans in answers:
        try:
            question = get_question(ans.question_number)
        except ValueError:
            continue
        marker = "✅ ДА" if is_yes(ans.answer) else "❌ НЕТ"
        sys_names = ", ".join(SYSTEMS[s].name for s in question.systems)
        lines.append(f"Вопрос #{question.number}: {question.text}")
        lines.append(f"   Ответ:    {marker}")
        lines.append(f"   Системы:  {sys_names}")
        lines.append("-" * 70)

    # ─── Дисклеймер ────────────────────────────────────────
    lines.append("")
    lines.append("═" * 70)
    lines.append(DISCLAIMER)
    lines.append("═" * 70)

    return "\n".join(lines).encode("utf-8")


# ════════════════════════════════════════════════════════════
#  Генерация Excel
# ════════════════════════════════════════════════════════════
def generate_excel_report(
    user: UserSnapshot,
    answers: List[AnswerInput],
    result: TestResult,
) -> bytes:
    """
    Excel-отчёт для CRM. Возвращает bytes (.xlsx).

    Структура:
      - Лист "Заявка": шапка с данными пользователя + результаты по системам
      - Лист "Ответы": все ответы построчно
    """
    # Импортируем тут, чтобы модуль работал даже если openpyxl не установлен
    # (тесты на TXT-генерацию пройдут).
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ─── Стили ──────────────────────────────────────────────
    header_fill = PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid")
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    border_thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    fill_good = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    fill_warning = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    fill_critical = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    # Подсветка блока реферера (мягкий жёлтый) и блока «холодного» (мягкий синий)
    fill_referrer = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    fill_cold = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

    status_to_fill = {
        "good": fill_good,
        "warning": fill_warning,
        "critical": fill_critical,
    }

    # ════════════════════════════════════════════════════════
    #  ЛИСТ 1: "Заявка"
    # ════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Заявка"

    # Заголовок
    ws["A1"] = f"WELLNESS TEST — Заявка #{user.lead_number}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:C1")

    # ─── Блок «Контактные данные» ──────────────────────────
    row = 3
    ws.cell(row=row, column=1, value="Поле").fill = header_fill
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=2, value="Значение").fill = header_fill
    ws.cell(row=row, column=2).font = header_font

    contact_data = [
        ("Дата", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("ФИО", user.full_name),
        ("Телефон", user.phone),
        ("Платформа", user.platform),
        ("User ID", str(user.platform_user_id)),
        ("Username", f"@{user.platform_username}" if user.platform_username else "—"),
    ]

    for label, value in contact_data:
        row += 1
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)

    # ─── Блок «Реферал» (расширенный или пометка «Зашёл напрямую») ──
    row += 1  # пустая строка-разделитель
    if user.referrer_name:
        # Реальный реферер: показываем все доступные контакты
        referrer_data = [
            ("Реферер ФИО", user.referrer_name),
            ("Реферер тел.", user.referrer_phone or "—"),
        ]
        if user.referrer_email:
            referrer_data.append(("Реферер email", user.referrer_email))
        if user.referrer_tg_username:
            referrer_data.append(("Реферер TG", f"@{user.referrer_tg_username}"))
        elif user.referrer_tg_user_id:
            referrer_data.append(("Реферер TG ID", str(user.referrer_tg_user_id)))
        if user.referrer_vk_user_id:
            referrer_data.append(
                ("Реферер VK", f"https://vk.com/id{user.referrer_vk_user_id}")
            )
        if user.referrer_max_user_id:
            referrer_data.append(("Реферер Max ID", str(user.referrer_max_user_id)))

        for label, value in referrer_data:
            row += 1
            cell_l = ws.cell(row=row, column=1, value=label)
            cell_l.font = Font(bold=True)
            cell_l.fill = fill_referrer
            cell_v = ws.cell(row=row, column=2, value=value)
            cell_v.fill = fill_referrer
    else:
        # «Холодный» лид — реферер = дистрибьютор или None
        row += 1
        cell_l = ws.cell(row=row, column=1, value="Реферал")
        cell_l.font = Font(bold=True)
        cell_l.fill = fill_cold
        cell_v = ws.cell(row=row, column=2, value="🌐 Зашёл напрямую (без реф-ссылки)")
        cell_v.fill = fill_cold

    # ─── Блок «Результаты по системам» ─────────────────────
    row += 2
    ws.cell(row=row, column=1, value="Результаты по системам").font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    row += 1
    headers = ["Система", "Баллы", "%", "Статус"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin

    for sys_result in result.systems:
        row += 1
        ws.cell(row=row, column=1, value=sys_result.system_name).border = border_thin
        ws.cell(row=row, column=2, value=f"{sys_result.score}/{sys_result.max_score}").border = border_thin
        ws.cell(row=row, column=3, value=f"{sys_result.percentage:.0f}%").border = border_thin
        status_cell = ws.cell(row=row, column=4, value=sys_result.status.label_ru)
        status_cell.fill = status_to_fill[sys_result.status.value]
        status_cell.border = border_thin

    # Ширина колонок на первом листе
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 25

    # ════════════════════════════════════════════════════════
    #  ЛИСТ 2: "Ответы"
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title="Ответы")

    headers2 = ["№", "Вопрос", "Ответ", "Системы"]
    for col_idx, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin

    row = 2
    for ans in answers:
        try:
            question = get_question(ans.question_number)
        except ValueError:
            continue
        sys_names = ", ".join(SYSTEMS[s].name for s in question.systems)
        ws2.cell(row=row, column=1, value=question.number).border = border_thin
        cell_q = ws2.cell(row=row, column=2, value=question.text)
        cell_q.alignment = wrap
        cell_q.border = border_thin
        ws2.cell(row=row, column=3, value="Да" if is_yes(ans.answer) else "Нет").border = border_thin
        ws2.cell(row=row, column=4, value=sys_names).border = border_thin
        ws2.row_dimensions[row].height = 35
        row += 1

    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 80
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 35

    # ─── Финальный дисклеймер на втором листе ──────────────
    row += 1
    ws2.cell(row=row, column=1, value=DISCLAIMER)
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws2.cell(row=row, column=1).alignment = wrap
    ws2.cell(row=row, column=1).font = Font(italic=True, size=9, color="C62828")
    ws2.row_dimensions[row].height = 60

    # ─── Сохраняем в bytes ─────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
